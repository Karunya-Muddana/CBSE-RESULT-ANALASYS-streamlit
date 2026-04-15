import re
import pandas as pd
import numpy as np
from io import BytesIO

SUBJECT_CODE_MAP = {
    "184": "English", "085": "Hindi(B)", "089": "Telugu", "018": "French",
    "041": "Maths(STD)", "086": "Science", "087": "Social", "049": "Painting",
    "165": "Computer Applications", "002": "Hindi(A)",
    "241": "Painting",   # Painting code used in Maths slot
}

# Codes that are valid 2nd-language subjects (position index 1)
LANG2_CODES = {"085", "089", "018", "165", "002"}

# Codes that are valid Maths-slot subjects (position index 2)
MATHS_SLOT_CODES = {"041", "241"}

GRADE_ORDER = ["A1","A2","B1","B2","C1","C2","D1","D2","E1","E2","F"]

GRADE_COLORS = {
    "A1":"1a9850","A2":"66bd63",
    "B1":"3182bd","B2":"6baed6",
    "C1":"fdae61","C2":"f46d43",
    "D1":"d73027","D2":"a50026",
    "E1":"969696","E2":"636363","F":"252525",
}

SUBJECTS     = ["English","Lang2","Maths","Science","Social"]
SUBJ_LABELS  = {"English":"English","Lang2":"2nd Language",
                "Maths":"Mathematics / Painting","Science":"Science","Social":"Social Science"}


def subject_name_from_code(code):
    if not code:
        return ""
    return SUBJECT_CODE_MAP.get(code, code)


def parse(lines_or_path):
    """
    Parse CBSE gazette text files.

    Supports two gazette formats:
      Format A — marks + grade letters on the second line:
                 e.g.  096 A1  089 B2  066 C1  ...
      Format B — marks only on the second line (no grade letters):
                 e.g.  096    089    066    063    081

    Auto-detects the format from the first marks line found.
    Grade letters are left empty ("") for Format B files.
    """
    if isinstance(lines_or_path, str):
        with open(lines_or_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.read().splitlines()
    else:
        lines = lines_or_path

    # ── Auto-detect format ──────────────────────────────────────────────────
    # Format A: marks line has "096 A1  089 B2" (digit + space + grade letter+digit)
    # Format B: marks line has "096    089    066" (digits only)
    # Use \b\d{2,3}\s+[A-F]\d\b to avoid false match on "026    C.B.S.E." in header.
    GRADE_PAT = re.compile(r"\b\d{2,3}\s+[A-F]\d\b")
    fmt_b = True   # default: marks-only (Format B)
    for line in lines:
        if re.match(r"\s*\d{8}", line):       # student header line — skip
            continue
        if GRADE_PAT.search(line):
            fmt_b = False   # grade letters present → Format A
            break
        if re.match(r"\s{10,}\d{2,3}\s", line):  # indented marks-only line
            fmt_b = True    # numbers only → Format B
            break

    records = []
    i = 0
    while i < len(lines):
        line = lines[i].rstrip("\r")
        m = re.match(
            r"(\d{8})\s+([MF])\s+(.+?)\s{3,}((?:\s*\d{3})+)\s+(PASS|FAIL|COMP|ESSEN|ABSE)",
            line
        )
        if m:
            roll, gender, name = m.group(1), m.group(2), m.group(3).strip()
            codes  = re.findall(r"\d{3}", m.group(4))
            result = m.group(5)

            # ── Advance to the marks line (skip blank lines only) ────────────
            j = i + 1
            while j < len(lines) and lines[j].strip() == "":
                j += 1

            marks_line = lines[j] if j < len(lines) else ""
            # Safety: don't consume another student's header line
            is_marks_line = (
                re.match(r"\s+\d{2,3}", marks_line)
                and not re.match(r"\d{8}", marks_line.strip())
            )

            if fmt_b:
                # Format B: extract bare numbers
                marks  = re.findall(r"\b(\d{2,3})\b", marks_line) if is_marks_line else []
                grades = [""] * len(marks)
            else:
                # Format A: extract (mark, grade) pairs
                pairs  = re.findall(r"(\d{1,3})\s+([A-F]\d?)", marks_line) if is_marks_line else []
                marks  = [p[0] for p in pairs]
                grades = [p[1] for p in pairs]

            # ── Identify subject roles ───────────────────────────────────────
            lang2_code = codes[1] if len(codes) > 1 else ""
            maths_code = codes[2] if len(codes) > 2 else ""

            rec = {
                "Roll":         roll,
                "Name":         name,
                "Gender":       gender,
                "Result":       result,
                "Lang2_Name":   subject_name_from_code(lang2_code),
                "Has_Painting": maths_code == "241",
            }

            for idx, subj in enumerate(SUBJECTS):
                code = codes[idx]  if idx < len(codes)  else ""
                rec[f"{subj}_Code"] = code
                rec[f"{subj}_Name"] = subject_name_from_code(code)
                rec[f"{subj}_M"]    = int(marks[idx])  if idx < len(marks)  else np.nan
                rec[f"{subj}_G"]    = grades[idx]       if idx < len(grades) else ""

            records.append(rec)
            i = j + 1
        else:
            i += 1

    df = pd.DataFrame(records)
    mark_cols = [f"{s}_M" for s in SUBJECTS]
    df["Total"] = df[mark_cols].sum(axis=1)
    df["Rank"]  = df["Total"].rank(method="min", ascending=False).astype(int)
    return df


def _safe_stats(s_ser):
    """Return (mean, median, max, min, std) safely even if series is empty."""
    if s_ser.empty:
        return 0.0, 0.0, 0, 0, 0.0
    return (
        round(float(s_ser.mean()), 1),
        round(float(s_ser.median()), 1),
        int(s_ser.max()),
        int(s_ser.min()),
        round(float(s_ser.std()), 1),
    )


def build_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()

    # ── Palette ────────────────────────────────────────────────────────────────
    C_DARK    = "1B3A5C"
    C_MID     = "2E75B6"
    C_LIGHT   = "D6E4F0"
    C_ACCENT  = "E8A838"
    C_GREEN   = "1A7A4A"
    C_GREEN2  = "D8F0E4"
    C_RED     = "C0392B"
    C_RED2    = "FAE0DD"
    C_WHITE   = "FFFFFF"
    C_GRAY    = "F4F6F9"
    C_ALTROW  = "EBF3FC"
    C_TEXT    = "1A1A2E"
    C_STRIPE  = "E2ECF7"
    C_PAINT   = "F4E4C1"   # warm highlight for Painting students

    thin  = Side(style="thin",   color="CBD5E0")
    thick = Side(style="medium", color="A0B4C8")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    gf = {g: PatternFill("solid", fgColor=c) for g, c in GRADE_COLORS.items()}

    def hdr(cell, text, bg=C_DARK, fg=C_WHITE, size=10, bold=True):
        cell.value = text
        cell.font  = Font(name="Calibri", bold=bold, color=fg, size=size)
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = brd

    def dat(cell, val, bold=False, center=True, bg=None, color=C_TEXT, size=10):
        cell.value = val
        cell.font  = Font(name="Calibri", bold=bold, color=color, size=size)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        cell.border = brd

    def section_hdr(ws, row, text, ncols, bg=C_MID):
        ws.row_dimensions[row].height = 28
        c = ws.cell(row, 1, text)
        c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=12)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = brd
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

    def title_row(ws, text, ncols, row=1, bg=C_DARK):
        ws.row_dimensions[row].height = 40
        c = ws.cell(row, 1, text)
        c.font = Font(name="Calibri", bold=True, color=C_WHITE, size=15)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

    def col_w(ws, col, w):
        ws.column_dimensions[get_column_letter(col)].width = w

    def color_scale(ws, ref):
        ws.conditional_formatting.add(ref, ColorScaleRule(
            start_type="min",  start_color="FA8072",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="63BE7B"
        ))

    # Count painting students for dashboard display
    painting_count = int(df.get("Has_Painting", pd.Series(False)).sum()) if "Has_Painting" in df.columns else 0

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — All Students
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "All Students"
    ws1.sheet_view.showGridLines = False

    cols1 = ["Rank","Roll","Name","Gender","Lang2_Name","Result",
             "English_Code","English_M","English_G",
             "Lang2_Code","Lang2_M","Lang2_G",
             "Maths_Code","Maths_M","Maths_G",
             "Science_Code","Science_M","Science_G",
             "Social_Code","Social_M","Social_G","Total"]
    hdrs1 = ["Rank","Roll No","Name","Gender","2nd Lang","Result",
             "Eng Code","Eng","Eng Gr",
             "L2 Code","Lang2","L2 Gr",
             "Math/Paint Code","Math/Paint","M/P Gr",
             "Sci Code","Science","Sci Gr",
             "Soc Code","Social","Soc Gr","Total/500"]
    n1 = len(cols1)

    title_row(ws1, "CBSE Class X Results 2026", n1)
    ws1.row_dimensions[2].height = 24
    for ci, lbl in enumerate(hdrs1, 1):
        hdr(ws1.cell(2, ci), lbl, bg=C_MID)

    df1 = df.sort_values("Total", ascending=False).reset_index(drop=True)
    for ri, (_, row) in enumerate(df1.iterrows(), 3):
        ws1.row_dimensions[ri].height = 18
        alt = ri % 2 == 0
        # Painting students get a warm background tint
        is_painter = row.get("Has_Painting", False)
        rbg = C_PAINT if is_painter else (C_ALTROW if alt else C_WHITE)
        for ci, col in enumerate(cols1, 1):
            v = row[col]
            if pd.isna(v): v = ""
            cell = ws1.cell(ri, ci)
            if col.endswith("_M") and v != "":
                cell.value = int(v)
            elif col == "Total" and v != "":
                cell.value = int(v)
            elif col == "Rank":
                medal = {1:"🥇",2:"🥈",3:"🥉"}.get(int(v),"")
                cell.value = f"{medal} {int(v)}" if medal else int(v)
            else:
                cell.value = v
            dat(cell, cell.value, bg=rbg, center=(ci != 3))
            if col.endswith("_G") and v in gf:
                cell.fill = gf[v]
                cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
            if col == "Total":
                cell.font = Font(name="Calibri", bold=True, color=C_DARK, size=11)
            # Flag Painting subject code cell
            if col == "Maths_Code" and v == "241":
                cell.fill = PatternFill("solid", fgColor="E8A838")
                cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)

    widths1 = [9,13,30,7,12,9,9,7,8,9,7,8,12,9,8,9,8,8,10,7,8,10]
    for i, w in enumerate(widths1, 1): col_w(ws1, i, w)
    ws1.freeze_panes = "C3"
    ws1.auto_filter.ref = f"A2:{get_column_letter(n1)}2"
    color_scale(ws1, f"{get_column_letter(n1)}3:{get_column_letter(n1)}{2+len(df)}")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Dashboard
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Dashboard")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_view.zoomScale = 90

    ws2.column_dimensions["A"].width = 22
    for c in range(2, 14):
        col_w(ws2, c, 13)

    title_row(ws2, "  CBSE Class X  —  Performance Dashboard", 12)

    ws2.row_dimensions[2].height = 5
    for c in range(1, 13):
        ws2.cell(2, c).fill = PatternFill("solid", fgColor=C_MID)

    kpi_config = [
        ("TOTAL STUDENTS",  len(df),                              C_MID,    1),
        ("PASS RATE",       "100%",                               C_GREEN,  3),
        ("CLASS AVERAGE",   f"{df['Total'].mean():.1f} / 500",    C_ACCENT, 5),
        ("HIGHEST SCORE",   f"{int(df['Total'].max())} / 500",    "7B2DBE", 7),
        ("BOYS / GIRLS",    f"{int((df.Gender=='M').sum())}  /  {int((df.Gender=='F').sum())}",
                                                                  "D63384", 9),
        ("PAINTING OPT.",   f"{painting_count} students",         "C07820", 11),
    ]

    ws2.row_dimensions[3].height = 18
    ws2.row_dimensions[4].height = 46
    ws2.row_dimensions[5].height = 6

    card_side = Side(style="thin", color="C8D8E8")
    for label, value, color, col in kpi_config:
        lc = ws2.cell(3, col, label)
        lc.font      = Font(name="Calibri", size=8, bold=True, color="7A90A4")
        lc.fill      = PatternFill("solid", fgColor="EFF4FA")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = Border(top=card_side, left=card_side, right=card_side)
        ws2.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)

        vc = ws2.cell(4, col, value)
        vc.font      = Font(name="Calibri", size=22, bold=True, color=color)
        vc.fill      = PatternFill("solid", fgColor=C_WHITE)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border    = Border(
            bottom=Side(style="thick", color=color),
            left=card_side, right=card_side,
        )
        ws2.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)

        gc = ws2.cell(5, col)
        gc.fill = PatternFill("solid", fgColor="E2ECF7")
        ws2.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)

    ws2.row_dimensions[6].height = 14

    # ── Subject Statistics ──────────────────────────────────────────────────
    section_hdr(ws2, 7, "  📊   Subject-wise Statistics", 8)
    ws2.row_dimensions[8].height = 22
    for ci, h in enumerate(["Subject","Average","Median","Highest","Lowest","Std Dev","A1+A2 %","Count"], 1):
        hdr(ws2.cell(8, ci), h, bg=C_DARK)

    for ri, subj in enumerate(SUBJECTS, 9):
        ws2.row_dimensions[ri].height = 22
        s_ser = pd.to_numeric(df[f"{subj}_M"], errors="coerce").dropna()
        a1a2  = df[f"{subj}_G"].isin(["A1","A2"]).sum()
        avg, med, hi, lo, std = _safe_stats(s_ser)
        # For Maths slot: show split between Maths takers and Painting takers
        if subj == "Maths" and painting_count > 0:
            maths_only = pd.to_numeric(
                df.loc[~df.get("Has_Painting", pd.Series(False)), f"{subj}_M"],
                errors="coerce"
            ).dropna()
            paint_only = pd.to_numeric(
                df.loc[df.get("Has_Painting", pd.Series(False)), f"{subj}_M"],
                errors="coerce"
            ).dropna()
            subj_label = f"Maths (n={len(maths_only)}) / Painting (n={len(paint_only)})"
        else:
            subj_label = SUBJ_LABELS[subj]

        vals = [subj_label, avg, med, hi, lo, std,
                f"{a1a2/len(df)*100:.0f}%" if len(df) else "—", int(a1a2)]
        rbg = C_ALTROW if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate(vals, 1):
            dat(ws2.cell(ri, ci), v, bg=rbg, bold=(ci == 1), center=(ci != 1))

    ws2.row_dimensions[14].height = 24
    ts = df["Total"]
    avg_t, med_t, hi_t, lo_t, std_t = _safe_stats(ts.dropna())
    for ci, v in enumerate(
        ["ALL SUBJECTS  (/500)", avg_t, med_t, hi_t, lo_t, std_t, "—", len(df)], 1
    ):
        hdr(ws2.cell(14, ci), v, bg=C_ACCENT, fg=C_DARK, size=10)

    color_scale(ws2, "B9:B13")

    ws2.row_dimensions[15].height = 14

    # ── Gender Comparison ────────────────────────────────────────────────────
    section_hdr(ws2, 16, "  👤   Gender Performance Comparison", 8)
    ws2.row_dimensions[17].height = 22
    for ci, h in enumerate(["Gender","English","Lang2","Math/Paint","Science","Social","Avg Total","Count"], 1):
        hdr(ws2.cell(17, ci), h, bg=C_DARK)

    for ri, (g, grp) in enumerate(df.groupby("Gender"), 18):
        ws2.row_dimensions[ri].height = 22
        avgs  = [round(pd.to_numeric(grp[f"{s}_M"], errors="coerce").mean(), 1) for s in SUBJECTS]
        is_f  = (g == "F")
        row_d = [("Female 👩" if is_f else "Male 👦")] + avgs + [round(grp["Total"].mean(), 1), len(grp)]
        rbg   = "FFF0F5" if is_f else "EFF6FF"
        for ci, v in enumerate(row_d, 1):
            dat(ws2.cell(ri, ci), v, bg=rbg, bold=(ci == 1), center=(ci != 1))

    ws2.row_dimensions[20].height = 14

    # ── 2nd Language Analysis ────────────────────────────────────────────────
    lang_list = list(df.groupby("Lang2_Name"))
    section_hdr(ws2, 21, "  🌐   2nd Language Group Analysis", 6)
    ws2.row_dimensions[22].height = 22
    for ci, h in enumerate(["Language","Students","% of Class","Avg Total","Avg Lang2","A1+A2 %"], 1):
        hdr(ws2.cell(22, ci), h, bg=C_DARK)

    for ri, (lang, grp) in enumerate(lang_list, 23):
        ws2.row_dimensions[ri].height = 22
        l2_avg   = pd.to_numeric(grp["Lang2_M"], errors="coerce").mean()
        a1a2_pct = f"{grp['Lang2_G'].isin(['A1','A2']).sum()/len(grp)*100:.0f}%"
        vals = [lang, len(grp), f"{len(grp)/len(df)*100:.1f}%",
                round(grp["Total"].mean(), 1), round(l2_avg, 1), a1a2_pct]
        rbg = C_ALTROW if ri % 2 == 0 else C_WHITE
        for ci, v in enumerate(vals, 1):
            dat(ws2.cell(ri, ci), v, bg=rbg, bold=(ci == 1), center=(ci != 1))

    # ── Painting vs Maths breakdown (only if painting students exist) ─────────
    if painting_count > 0:
        paint_row = 23 + len(lang_list) + 2
        section_hdr(ws2, paint_row, "  🎨   Maths vs Painting — Performance Comparison", 6)
        ws2.row_dimensions[paint_row + 1].height = 22
        for ci, h in enumerate(["Group","Students","Avg Marks","Median","Highest","Lowest"], 1):
            hdr(ws2.cell(paint_row + 1, ci), h, bg=C_DARK)

        groups = [
            ("Maths (041)",    ~df["Has_Painting"]),
            ("Painting (241)",  df["Has_Painting"]),
        ]
        for offset, (label, mask) in enumerate(groups):
            ri = paint_row + 2 + offset
            ws2.row_dimensions[ri].height = 22
            grp_ser = pd.to_numeric(df.loc[mask, "Maths_M"], errors="coerce").dropna()
            avg2, med2, hi2, lo2, _ = _safe_stats(grp_ser)
            row_vals = [label, int(mask.sum()), avg2, med2, hi2, lo2]
            rbg = C_PAINT if "Painting" in label else C_ALTROW
            for ci, v in enumerate(row_vals, 1):
                dat(ws2.cell(ri, ci), v, bg=rbg, bold=(ci == 1), center=(ci != 1))

    # ── Charts ───────────────────────────────────────────────────────────────
    _DC = 14
    _DR = 3
    ws2.cell(_DR,   _DC, "Subject")
    ws2.cell(_DR,   _DC+1, "Average")
    for ri, subj in enumerate(SUBJECTS, _DR+1):
        ws2.cell(ri, _DC,   SUBJ_LABELS[subj])
        ws2.cell(ri, _DC+1, round(pd.to_numeric(df[f"{subj}_M"], errors="coerce").mean(), 1))
    ws2.column_dimensions[get_column_letter(_DC)].hidden   = True
    ws2.column_dimensions[get_column_letter(_DC+1)].hidden = True

    bar = BarChart()
    bar.type  = "col"; bar.style = 10
    bar.title = "Class Average by Subject"
    bar.y_axis.title = "Average Marks"
    bar.width  = 17; bar.height = 12
    bar.y_axis.scaling.min = 50; bar.y_axis.scaling.max = 100
    bar.add_data(Reference(ws2, min_col=_DC+1, min_row=_DR, max_row=_DR+5), titles_from_data=True)
    bar.set_categories(Reference(ws2, min_col=_DC, min_row=_DR+1, max_row=_DR+5))
    bar.series[0].graphicalProperties.solidFill = C_MID
    ws2.add_chart(bar, "J3")

    _GD = _DR + 8
    ws2.cell(_GD, _DC, "Gender"); ws2.cell(_GD, _DC+1, "Count")
    for ri, (g, cnt) in enumerate(df["Gender"].value_counts().items(), _GD+1):
        ws2.cell(ri, _DC,   "Male" if g == "M" else "Female")
        ws2.cell(ri, _DC+1, int(cnt))
    pie = PieChart()
    pie.style = 10; pie.title = "Gender Distribution"
    pie.width = 14; pie.height = 10
    pie.add_data(Reference(ws2, min_col=_DC+1, min_row=_GD, max_row=_GD+2), titles_from_data=True)
    pie.set_categories(Reference(ws2, min_col=_DC, min_row=_GD+1, max_row=_GD+2))
    ws2.add_chart(pie, "J17")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Grade Distribution
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Grade Distribution")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 20
    for i in range(2, 16): col_w(ws3, i, 9)

    present = [g for g in GRADE_ORDER
               if any(df[f"{s}_G"].eq(g).any() for s in SUBJECTS)]
    n3 = 1 + len(present) + 2

    title_row(ws3, "Grade Distribution — Subject × Grade Matrix", n3)
    hdrs3 = ["Subject"] + present + ["Total","A1+A2 %"]
    ws3.row_dimensions[2].height = 22
    for ci, h in enumerate(hdrs3, 1):
        bg = GRADE_COLORS.get(h, C_MID)
        hdr(ws3.cell(2, ci), h, bg=bg, fg="FFFFFF" if h in GRADE_COLORS else C_WHITE)

    for ri, subj in enumerate(SUBJECTS, 3):
        ws3.row_dimensions[ri].height = 20
        label = SUBJ_LABELS[subj]
        if subj == "Maths" and painting_count > 0:
            label = f"Maths/Painting ({painting_count}🎨)"
        ws3.cell(ri, 1, label).font = Font(name="Calibri", bold=True, size=10, color=C_DARK)
        ws3.cell(ri, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws3.cell(ri, 1).border = brd
        ws3.cell(ri, 1).fill   = PatternFill("solid", fgColor=C_ALTROW if ri%2==0 else C_WHITE)
        total_n = 0; a1a2_n = 0
        for ci, g in enumerate(present, 2):
            cnt  = int(df[f"{subj}_G"].eq(g).sum())
            cell = ws3.cell(ri, ci, cnt if cnt > 0 else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd
            if cnt > 0 and g in gf:
                cell.fill = gf[g]
                cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
            else:
                cell.fill = PatternFill("solid", fgColor=C_ALTROW if ri%2==0 else C_WHITE)
                cell.font = Font(name="Calibri", color="BBBBBB", size=10)
            total_n += cnt
            if g in ("A1","A2"): a1a2_n += cnt
        dat(ws3.cell(ri, len(present)+2), total_n, bold=True, bg=C_LIGHT)
        pct = f"{a1a2_n/total_n*100:.0f}%" if total_n else ""
        dat(ws3.cell(ri, len(present)+3), pct, bold=True,
            bg=C_GREEN2 if a1a2_n/max(total_n,1) >= 0.5 else C_RED2)

    ws3.row_dimensions[8].height = 22
    for ci, g in enumerate(present, 2):
        cnt = sum(int(df[f"{s}_G"].eq(g).sum()) for s in SUBJECTS)
        hdr(ws3.cell(8, ci), cnt, bg=C_DARK)
    hdr(ws3.cell(8, 1), "CLASS TOTAL", bg=C_DARK)
    hdr(ws3.cell(8, len(present)+2), len(df)*5, bg=C_DARK)

    _GD2 = 10
    ws3.cell(_GD2, 1, "Subject")
    use_g = present[:7]
    for ci, g in enumerate(use_g, 2): ws3.cell(_GD2, ci, g)
    for ri, subj in enumerate(SUBJECTS, _GD2+1):
        ws3.cell(ri, 1, SUBJ_LABELS[subj])
        for ci, g in enumerate(use_g, 2):
            ws3.cell(ri, ci, int(df[f"{subj}_G"].eq(g).sum()))

    bar2 = BarChart(); bar2.type="col"; bar2.grouping="stacked"; bar2.style=10
    bar2.title = "Grade Distribution per Subject"
    bar2.y_axis.title = "Students"; bar2.width = 18; bar2.height = 12
    bar2.add_data(Reference(ws3, min_col=2, min_row=_GD2, max_col=len(use_g)+1, max_row=_GD2+5),
                  titles_from_data=True)
    bar2.set_categories(Reference(ws3, min_col=1, min_row=_GD2+1, max_row=_GD2+5))
    grade_fills = ["1a9850","66bd63","3182bd","6baed6","fdae61","f46d43","d73027"]
    for i, s in enumerate(bar2.series):
        if i < len(grade_fills): s.graphicalProperties.solidFill = grade_fills[i]
    ws3.add_chart(bar2, "A16")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Rank List
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Rank List")
    ws4.sheet_view.showGridLines = False

    cols4 = ["Rank","Roll","Name","Gender","Lang2_Name",
             "English_Code","English_M",
             "Lang2_Code","Lang2_M",
             "Maths_Code","Maths_M",
             "Science_Code","Science_M",
             "Social_Code","Social_M","Total"]
    hdrs4 = ["Rank","Roll No","Name","Gender","2nd Lang",
             "Eng Code","English",
             "L2 Code","Lang2",
             "Math/Paint Code","Math/Paint",
             "Sci Code","Science",
             "Soc Code","Social","Total /500"]
    n4 = len(cols4)

    title_row(ws4, "Overall Rank List — CBSE Class X 2026", n4)
    ws4.row_dimensions[2].height = 22
    for ci, h in enumerate(hdrs4, 1): hdr(ws4.cell(2, ci), h, bg=C_MID)

    df4 = df.sort_values("Total", ascending=False).reset_index(drop=True)
    for ri, (_, row) in enumerate(df4.iterrows(), 3):
        ws4.row_dimensions[ri].height = 18
        rank     = int(row["Rank"])
        is_paint = row.get("Has_Painting", False)
        rbg = {"1":"FFF9E6","2":"F5F5F5","3":"FFF0E6"}.get(
            str(rank), C_PAINT if is_paint else (C_ALTROW if ri%2==0 else C_WHITE)
        )
        for ci, col in enumerate(cols4, 1):
            v = row[col]
            if pd.isna(v): v = ""
            cell = ws4.cell(ri, ci)
            if col == "Rank":
                medal = {1:"🥇",2:"🥈",3:"🥉"}.get(rank,"")
                cell.value = f"{medal} {rank}" if medal else rank
            elif col in ("Total",) + tuple(f"{s}_M" for s in SUBJECTS) and v != "":
                cell.value = int(v)
            else:
                cell.value = v
            dat(cell, cell.value, bg=rbg, bold=(col=="Total" or rank<=3), center=(ci!=3))
            if col == "Total":
                cell.font = Font(name="Calibri", bold=True, color=C_DARK, size=11)
            # Highlight Painting code cell in amber
            if col == "Maths_Code" and v == "241":
                cell.fill = PatternFill("solid", fgColor="E8A838")
                cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)

    widths4 = [9,13,30,8,12,9,10,9,10,13,11,9,10,9,10,11]
    for i, w in enumerate(widths4, 1): col_w(ws4, i, w)
    ws4.freeze_panes = "A3"
    ws4.auto_filter.ref = f"A2:{get_column_letter(n4)}2"
    color_scale(ws4, f"{get_column_letter(n4)}3:{get_column_letter(n4)}{2+len(df)}")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEETS 5–9 — Per-subject rank lists
    # ══════════════════════════════════════════════════════════════════════════
    for subj in SUBJECTS:
        label = SUBJ_LABELS[subj]
        ws    = wb.create_sheet(label[:12])
        ws.sheet_view.showGridLines = False

        scols = ["Roll","Name","Gender","Lang2_Name",f"{subj}_Code",f"{subj}_M",f"{subj}_G","Total"]
        shdrs = ["Roll No","Name","Gender","2nd Lang","Sub Code","Marks","Grade","Total"]
        ns    = len(scols) + 1

        title_row(ws, f"Subject Rank — {label}", ns)
        ws.row_dimensions[2].height = 22
        hdr(ws.cell(2,1), "Rank", bg=C_MID)
        for ci, h in enumerate(shdrs, 2): hdr(ws.cell(2, ci), h, bg=C_MID)

        dfs = df.sort_values(f"{subj}_M", ascending=False).reset_index(drop=True)
        for ri, (_, row) in enumerate(dfs.iterrows(), 3):
            ws.row_dimensions[ri].height = 18
            sr       = ri - 2
            alt      = ri % 2 == 0
            is_paint = row.get("Has_Painting", False) and subj == "Maths"
            rbg = {"1":"FFF9E6","2":"F5F5F5","3":"FFF0E6"}.get(
                str(sr), C_PAINT if is_paint else (C_ALTROW if alt else C_WHITE)
            )
            medal = {1:"🥇",2:"🥈",3:"🥉"}.get(sr,"")
            c = ws.cell(ri, 1, f"{medal} {sr}" if medal else sr)
            dat(c, c.value, bg=rbg, bold=(sr<=3))
            for ci, col in enumerate(scols, 2):
                v = row[col]
                if pd.isna(v): v = ""
                cell = ws.cell(ri, ci)
                cell.value = int(v) if isinstance(v, float) and col.endswith("_M") else \
                             (int(v) if col=="Total" and v != "" else v)
                dat(cell, cell.value, bg=rbg, center=(ci!=3))
                if col == f"{subj}_G" and v in gf:
                    cell.fill = gf[v]
                    cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
                if col == f"{subj}_Code" and v == "241":
                    cell.fill = PatternFill("solid", fgColor="E8A838")
                    cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)

        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 13
        ws.column_dimensions["C"].width = 30
        for i in range(4, ns+1): col_w(ws, i, 11)
        ws.freeze_panes = "A3"

        # ── Summary stats block ───────────────────────────────────────────────
        sr2 = len(df) + 4
        s_s = pd.to_numeric(df[f"{subj}_M"], errors="coerce").dropna()
        a1a2 = df[f"{subj}_G"].isin(["A1","A2"]).sum()
        avg, med, hi, lo, std = _safe_stats(s_s)

        title_row(ws, f"{label} — Summary Statistics", ns, row=sr2, bg=C_MID)
        stats_data = [
            ("Average Marks",  avg),
            ("Median",         med),
            ("Highest",        hi),
            ("Lowest",         lo),
            ("Std Deviation",  std),
            ("A1 + A2 Count",  int(a1a2)),
            ("A1 + A2 %",      f"{a1a2/len(df)*100:.0f}%" if len(df) else "—"),
        ]

        # For Maths slot — append Painting sub-breakdown
        if subj == "Maths" and painting_count > 0:
            maths_only = pd.to_numeric(
                df.loc[~df["Has_Painting"], "Maths_M"], errors="coerce"
            ).dropna()
            paint_only = pd.to_numeric(
                df.loc[df["Has_Painting"],  "Maths_M"], errors="coerce"
            ).dropna()
            m_avg, _, m_hi, m_lo, _ = _safe_stats(maths_only)
            p_avg, _, p_hi, p_lo, _ = _safe_stats(paint_only)
            stats_data += [
                ("— Maths (041) takers",   int(len(maths_only))),
                ("  Avg (Maths only)",      m_avg),
                ("— Painting (241) takers", int(len(paint_only))),
                ("  Avg (Painting only)",   p_avg),
            ]

        for i, (lbl2, v) in enumerate(stats_data, sr2+1):
            lc = ws.cell(i, 1, lbl2)
            lc.font      = Font(name="Calibri", bold=True, size=10)
            lc.alignment = Alignment(horizontal="left", vertical="center")
            lc.fill      = PatternFill("solid", fgColor=C_ALTROW if i%2==0 else C_WHITE)
            lc.border    = brd
            vc = ws.cell(i, 2, v)
            vc.font      = Font(name="Calibri", bold=True, size=10, color=C_MID)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.fill      = PatternFill("solid", fgColor=C_ALTROW if i%2==0 else C_WHITE)
            vc.border    = brd

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 10 — Needs Attention
    # ══════════════════════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("Needs Attention")
    ws10.sheet_view.showGridLines = False

    avg_t = df["Total"].mean()
    std_t = df["Total"].std()
    thr   = avg_t - std_t
    flag  = (
        df[["English_M","Lang2_M","Maths_M","Science_M","Social_M"]].lt(60).any(axis=1) |
        df["Total"].lt(thr)
    )
    df_na = df[flag].sort_values("Total").reset_index(drop=True)

    na_cols = ["Rank","Roll","Name","Gender",
               "English_Code","English_M",
               "Lang2_Code","Lang2_M",
               "Maths_Code","Maths_M",
               "Science_Code","Science_M",
               "Social_Code","Social_M","Total"]
    na_hdrs = ["Rank","Roll No","Name","Gender",
               "Eng Code","English",
               "L2 Code","Lang2",
               "Math/Paint Code","Math/Paint",
               "Sci Code","Science",
               "Soc Code","Social","Total"]
    n10 = len(na_cols)

    title_row(ws10, f"⚠  Needs Attention — {len(df_na)} Students", n10, bg=C_RED)
    sub = ws10.cell(2, 1,
          f"Criteria: any subject < 60  OR  total < {thr:.0f}  "
          f"(class avg {avg_t:.1f} − 1 SD {std_t:.1f})")
    sub.font = Font(name="Calibri", italic=True, size=9, color="888888")
    ws10.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n10)
    ws10.row_dimensions[2].height = 16
    ws10.row_dimensions[3].height = 22
    for ci, h in enumerate(na_hdrs, 1): hdr(ws10.cell(3, ci), h, bg="8B0000")

    for ri, (_, row) in enumerate(df_na.iterrows(), 4):
        ws10.row_dimensions[ri].height = 18
        is_paint = row.get("Has_Painting", False)
        for ci, col in enumerate(na_cols, 1):
            v = row[col]
            if pd.isna(v): v = ""
            cell = ws10.cell(ri, ci)
            cell.value = int(v) if isinstance(v, float) and v != "" else v
            is_low = col.endswith("_M") and isinstance(v, (int,float)) and v < 60
            bg = "FFD5D5" if is_low else (C_PAINT if is_paint else (C_ALTROW if ri%2==0 else C_WHITE))
            dat(cell, cell.value, bg=bg, center=(ci!=3))
            if is_low:
                cell.font = Font(name="Calibri", bold=True, color=C_RED, size=10)
            if col == "Maths_Code" and v == "241":
                cell.fill = PatternFill("solid", fgColor="E8A838")
                cell.font = Font(name="Calibri", bold=True, color=C_WHITE, size=10)

    ws10.column_dimensions["A"].width = 7
    ws10.column_dimensions["B"].width = 13
    ws10.column_dimensions["C"].width = 30
    ws10.column_dimensions["D"].width = 8
    widths10 = [7,13,30,8,9,10,9,10,13,11,9,10,9,10,11]
    for i, w in enumerate(widths10, 1):
        col_w(ws10, i, w)
    ws10.freeze_panes = "A4"
    if len(df_na) > 0:
        color_scale(ws10, f"{get_column_letter(n10)}4:{get_column_letter(n10)}{3+len(df_na)}")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()